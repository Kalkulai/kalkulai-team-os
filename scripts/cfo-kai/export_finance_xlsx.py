#!/usr/bin/env python3
"""Build the CFO-Kai finance workbook from a FinanceData snapshot."""
from __future__ import annotations

import argparse
import json
import subprocess
import sys
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parent))

from cfo_common import (  # noqa: E402
    DEFAULT_BASE,
    drive_link_for,
    load_or_fetch,
    resolve_output_dir,
    safe_month_label,
)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover - exercised on misconfigured bridge hosts.
    raise SystemExit("ERROR: openpyxl is required for export_finance_xlsx.py") from exc


BLUE = "0000FF"
BLACK = "000000"
WHITE = "FFFFFF"
YELLOW = "FFFF00"
HEADER_FILL = "1F4E78"
SUBTLE_FILL = "D9EAF7"
GRID = "B7C9D6"
CURRENCY_FORMAT = '#,##0 €;(#,##0 €);-'
NUMBER_FORMAT = '#,##0.0;-#,##0.0;-'
ERROR_VALUES = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!"}

DEFAULT_RECALC = (
    Path.home()
    / ".claude/plugins/cache/anthropic-agent-skills/document-skills/da20c92503b2/"
    / "skills/xlsx/scripts/recalc.py"
)


def build_workbook(
    data: dict[str, Any],
    output_dir: str | Path | None = None,
    recalc: bool = True,
    now_label: str | None = None,
    recalc_script: str | Path | None = None,
) -> Path:
    month = safe_month_label(now_label)
    out_dir = resolve_output_dir(output_dir)
    path = out_dir / f"cfo-kai-finance-{month}.xlsx"

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    inputs = wb.create_sheet("Inputs")
    costs = wb.create_sheet("Costs")
    forecast = wb.create_sheet("Forecast")
    assumptions = wb.create_sheet("Assumptions")

    _build_inputs(inputs, data)
    _build_summary(summary, data)
    _build_costs(costs, data)
    _build_forecast(forecast, data)
    _build_assumptions(assumptions, data)

    for ws in wb.worksheets:
        _finish_sheet(ws)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(path)

    if recalc:
        run_recalc(path, Path(recalc_script) if recalc_script else DEFAULT_RECALC)
    errors = scan_formula_errors(path)
    if errors:
        raise RuntimeError(f"formula errors found after export: {json.dumps(errors, sort_keys=True)}")
    return path


def _title(ws: Any, title: str, subtitle: str) -> None:
    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=HEADER_FILL)
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Arial", italic=True, color=BLACK)
    ws.merge_cells("A1:D1")
    ws.merge_cells("A2:D2")


def _build_inputs(ws: Any, data: dict[str, Any]) -> None:
    _title(ws, "CFO-Kai Finance Inputs", "Blue font = source snapshot input; yellow fill = assumption/source context.")
    rows = [
        ("Metric", "Value", "Unit", "Source"),
        ("Cash on hand", data["cash_on_hand_eur"], "EUR", "GET /api/finance"),
        ("Monthly burn actual", data["monthly_burn"]["actual_eur"], "EUR/month", "GET /api/finance"),
        ("Monthly burn plan", data["monthly_burn"]["plan_eur"], "EUR/month", "GET /api/finance"),
        ("Runway", data["runway_months"], "months", "GET /api/finance"),
        ("Break-even", data["break_even_label"], "", "FinanceData.break_even_label"),
        ("As of", data["as_of"], "", "FinanceData.as_of"),
        ("Generated at", data.get("generated_at", ""), "", "FinanceData.generated_at"),
    ]
    for row in rows:
        ws.append(row)
    _style_table(ws, 3, len(rows) + 2)
    for row in range(4, len(rows) + 3):
        ws.cell(row, 2).font = Font(name="Arial", color=BLUE)
    for row in (8, 9, 10):
        for col in range(1, 5):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=YELLOW)
    for row in (4, 5, 6):
        ws.cell(row, 2).number_format = CURRENCY_FORMAT
    ws.cell(7, 2).number_format = NUMBER_FORMAT


def _build_summary(ws: Any, data: dict[str, Any]) -> None:
    _title(ws, "CFO-Kai Finance Export", "Formula-driven summary matched to the dashboard FinanceData contract.")
    rows = [
        ("KPI", "Value", "Unit", "Formula / note"),
        ("Cash on hand", "=Inputs!B4", "EUR", "Dashboard KPI"),
        ("Runway", "=Inputs!B7", "months", "Dashboard KPI"),
        ("Monthly burn actual", "=Inputs!B5", "EUR/month", "Dashboard KPI"),
        ("Monthly burn plan", "=Inputs!B6", "EUR/month", "Plan baseline"),
        ("Burn delta", "=B6-B7", "EUR/month", "actual - plan"),
        ("Total cost lines", f"=SUM(Costs!B2:B{len(data['cost_lines']) + 1})", "EUR/month", "Costs sheet total"),
        ("Break-even", "=Inputs!B8", "", "Dashboard KPI"),
    ]
    for row in rows:
        ws.append(row)
    _style_table(ws, 3, len(rows) + 2)
    for row in range(4, 10):
        ws.cell(row, 2).font = Font(name="Arial", color=BLACK)
    for row in (4, 6, 7, 8, 9):
        ws.cell(row, 2).number_format = CURRENCY_FORMAT
    ws.cell(5, 2).number_format = NUMBER_FORMAT


def _build_costs(ws: Any, data: dict[str, Any]) -> None:
    ws.append(["Label", "Amount", "Fixed", "Paid by", "Note"])
    for line in data["cost_lines"]:
        ws.append([
            line.get("label", ""),
            line.get("amount_eur", 0),
            "yes" if line.get("fixed") else "no",
            line.get("paid_by", ""),
            line.get("note", ""),
        ])
    total_row = len(data["cost_lines"]) + 2
    ws.append(["Total", f"=SUM(B2:B{total_row - 1})", "", "", "Formula total"])
    _style_table(ws, 1, total_row)
    for row in range(2, total_row):
        for col in range(1, 6):
            ws.cell(row, col).font = Font(name="Arial", color=BLUE)
        ws.cell(row, 2).number_format = CURRENCY_FORMAT
    ws.cell(total_row, 1).font = Font(name="Arial", bold=True, color=BLACK)
    ws.cell(total_row, 2).font = Font(name="Arial", bold=True, color=BLACK)
    ws.cell(total_row, 2).number_format = CURRENCY_FORMAT


def _build_forecast(ws: Any, data: dict[str, Any]) -> None:
    ws.append(["Month", "Cash", "Burn", "Cash delta"])
    for idx, point in enumerate(data["forecast_6m"], start=2):
        delta_formula = "-" if idx == 2 else f"=B{idx}-B{idx - 1}"
        ws.append([point.get("month", ""), point.get("cash_eur", 0), point.get("burn_eur", 0), delta_formula])
    total_row = len(data["forecast_6m"]) + 2
    ws.append(["Total burn", "", f"=SUM(C2:C{total_row - 1})", ""])
    _style_table(ws, 1, total_row)
    for row in range(2, total_row):
        for col in range(1, 4):
            ws.cell(row, col).font = Font(name="Arial", color=BLUE)
        ws.cell(row, 2).number_format = CURRENCY_FORMAT
        ws.cell(row, 3).number_format = CURRENCY_FORMAT
        if row > 2:
            ws.cell(row, 4).font = Font(name="Arial", color=BLACK)
            ws.cell(row, 4).number_format = CURRENCY_FORMAT
    ws.cell(total_row, 3).font = Font(name="Arial", bold=True, color=BLACK)
    ws.cell(total_row, 3).number_format = CURRENCY_FORMAT


def _build_assumptions(ws: Any, data: dict[str, Any]) -> None:
    _title(ws, "Assumptions and Sources", "Yellow cells require review when the source snapshot changes.")
    ws.append(["Topic", "Status / Value", "Note"])
    ws.append(["Snapshot", data["as_of"], "Source: GET /api/finance"])
    ws.append(["Break-even", data["break_even_label"], "Source: FinanceData.break_even_label"])
    for row in data["pilot_health"]:
        ws.append([
            row.get("name", ""),
            row.get("status", ""),
            row.get("note", ""),
        ])
    _style_table(ws, 3, ws.max_row)
    for row in range(4, ws.max_row + 1):
        for col in range(1, 4):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=YELLOW)
            ws.cell(row, col).font = Font(name="Arial", color=BLUE)


def _style_table(ws: Any, header_row: int, max_row: int) -> None:
    thin = Side(style="thin", color=GRID)
    for cell in ws[header_row]:
        cell.font = Font(name="Arial", bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=header_row, max_row=max_row):
        for cell in row:
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if cell.row > header_row and cell.row % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=SUBTLE_FILL)


def _finish_sheet(ws: Any) -> None:
    ws.freeze_panes = "A4" if ws.title in {"Summary", "Inputs", "Assumptions"} else "A2"
    for column_cells in ws.columns:
        width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(width + 2, 12), 42)
    for row in ws.iter_rows():
        for cell in row:
            if cell.font is None or cell.font.name is None:
                cell.font = Font(name="Arial", color=BLACK)


def run_recalc(path: Path, recalc_script: Path) -> None:
    if not recalc_script.exists():
        raise RuntimeError(f"recalc.py not found: {recalc_script}")
    result = subprocess.run(
        [sys.executable, str(recalc_script), str(path), "45"],
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=False,
    )
    if result.returncode != 0:
        raise RuntimeError(f"recalc.py failed: {result.stderr or result.stdout}")
    try:
        payload = json.loads(result.stdout)
    except json.JSONDecodeError:
        return
    if payload.get("status") == "errors_found" or payload.get("total_errors", 0):
        raise RuntimeError(f"recalc.py found formula errors: {result.stdout}")


def scan_formula_errors(path: str | Path) -> dict[str, list[str]]:
    wb = load_workbook(path, data_only=False)
    errors: dict[str, list[str]] = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str):
                    for err in ERROR_VALUES:
                        if err in value:
                            errors.setdefault(err, []).append(f"{ws.title}!{cell.coordinate}")
    return errors


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="Export /api/finance to a polished CFO-Kai XLSX workbook")
    ap.add_argument("--file", help="Local FinanceData JSON. If omitted, GET /api/finance is used.")
    ap.add_argument("--scenario", choices=["current", "exist"], default=None)
    ap.add_argument("--base", default=DEFAULT_BASE)
    ap.add_argument("--secret", default="")
    ap.add_argument("--output-dir", default=None)
    ap.add_argument("--month", default=None, help="Output month label YYYY-MM")
    ap.add_argument("--skip-recalc", action="store_true", help="Only for local tests/debugging.")
    ap.add_argument("--recalc-script", default=str(DEFAULT_RECALC))
    return ap.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    if not args.secret:
        args.secret = __import__("os").environ.get("DASHBOARD_API_SECRET", "")
    try:
        data = load_or_fetch(args)
        path = build_workbook(
            data,
            output_dir=args.output_dir,
            recalc=not args.skip_recalc,
            now_label=args.month,
            recalc_script=args.recalc_script,
        )
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    print(f"OK: xlsx={path} link={drive_link_for(path)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
